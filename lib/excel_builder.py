"""Excel出力 — 発注シミュレーター生成"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ── デザイン定数 ──────────────────────────────
FONT_NAME = "Arial"
MAIN_COLOR = "1B4F72"
HEADER_BG = "D6EAF8"
THIN_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
HEADER_FONT = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=MAIN_COLOR, end_color=MAIN_COLOR, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
RED_FILL = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
GREEN_FILL = PatternFill(start_color="ABEBC6", end_color="ABEBC6", fill_type="solid")
BLUE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="0000FF")
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11, color=MAIN_COLOR)
SECTION_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
NUM_FMT_INT = '#,##0'
NUM_FMT_YEN = '¥#,##0'
NUM_FMT_DEC = '#,##0.0'

# 発注残セルのスタイル
BACKORDER_FONT = Font(name=FONT_NAME, size=10, color="0000FF")
BACKORDER_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

# 詳細シートカラム定義 (A-S = 19列)
DETAIL_HEADERS = [
    "SKUコード", "商品名", "区分", "現在庫合計", "カインズ", "RSL", "FBA",
    "発注残", "過去12ヶ月販売", "月平均(6ヶ月)", "月間予測需要", "安全在庫",
    "目標在庫", "発注推奨数(個)", "発注ケース数", "発注金額概算",
    "ケース入数", "仕入単価", "緊急度",
]
COL_WIDTHS = [18, 30, 8, 10, 8, 8, 8, 8, 12, 10, 10, 10, 10, 12, 10, 12, 8, 8, 8]


def _apply_border(ws, row, max_col):
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


def _apply_borders_range(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        _apply_border(ws, r, max_col)


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ──────────────────────────────────────────────
# メインビルド関数
# ──────────────────────────────────────────────

def build_simulator_excel(
    output_path: Path,
    sku_data: list[dict],
    maker_type: str,  # "flexi" or "petzpark"
    seasonal_indices: pd.Series = None,
    params: dict = None,
):
    """発注シミュレーター Excel を生成

    Args:
        output_path: 出力ファイルパス
        sku_data: SKUごとのデータ辞書リスト
        maker_type: "flexi" or "petzpark"
        seasonal_indices: 月別季節指数 (flexi版のみ)
        params: {"lt": int, "oc": int}
    """
    if params is None:
        params = {"lt": 75, "oc": 60} if maker_type == "flexi" else {"lt": 14, "oc": 30}

    wb = Workbook()

    # ── シート2: 発注詳細（先に作成、数式参照のため）
    ws_detail = wb.active
    ws_detail.title = "発注詳細"
    subtotal_rows = _build_detail_sheet(ws_detail, sku_data, maker_type, params)

    # ── シート1: シミュレーション（先頭に移動）
    ws_sim = wb.create_sheet("シミュレーション", 0)
    _build_simulation_sheet(ws_sim, maker_type, seasonal_indices, subtotal_rows, params)

    # ── シート3: 季節指数（flexi版のみ）
    if maker_type == "flexi" and seasonal_indices is not None:
        ws_si = wb.create_sheet("季節指数")
        _build_seasonal_sheet(ws_si, seasonal_indices)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ──────────────────────────────────────────────
# シート1: シミュレーション
# ──────────────────────────────────────────────

def _build_simulation_sheet(ws, maker_type, seasonal_indices, subtotal_rows, params):
    ws.sheet_properties.tabColor = MAIN_COLOR

    title = "flexi 発注シミュレーター" if maker_type == "flexi" else "Petz Park 発注シミュレーター"
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(name=FONT_NAME, bold=True, size=16, color=MAIN_COLOR)

    ws["A3"] = "発注係数"
    ws["A3"].font = Font(name=FONT_NAME, bold=True, size=11)
    ws["B5"] = 1.0
    ws["B5"].font = BLUE_FONT
    ws["B5"].fill = YELLOW_FILL
    ws["B5"].number_format = '0.0'
    ws["B5"].alignment = Alignment(horizontal="center")
    ws["C5"] = "← 0.5〜1.5 で調整"
    ws["C5"].font = Font(name=FONT_NAME, size=9, color="666666")

    # サマリー
    row = 7
    ws.cell(row=row, column=1, value="■ サマリー").font = SECTION_FONT

    # subtotal_rows = {"main_subtotal": int, "sub_subtotal": int, "grand_total": int}
    main_sub = subtotal_rows.get("main_subtotal")
    grand_total = subtotal_rows.get("grand_total")

    labels_and_refs = []
    if maker_type == "flexi":
        main_label = "本体"
    else:
        main_label = "通常サイズ"

    labels_and_refs = [
        (f"{main_label} 合計発注数（個）", "N", main_sub, NUM_FMT_INT),
        (f"{main_label} 合計発注ケース数", "O", main_sub, NUM_FMT_INT),
        (f"{main_label} 合計発注金額（税抜概算）", "P", main_sub, NUM_FMT_YEN),
        ("全体 合計発注金額", "P", grand_total, NUM_FMT_YEN),
    ]
    for i, (label, col, ref_row, fmt) in enumerate(labels_and_refs):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT_NAME, size=10)
        cell = ws.cell(row=r, column=3)
        if ref_row:
            cell.value = f"=発注詳細!{col}{ref_row}"
        else:
            cell.value = 0
        cell.number_format = fmt
        cell.font = Font(name=FONT_NAME, bold=True, size=11)

    # 季節指数テーブル (flexi版のみ)
    if maker_type == "flexi" and seasonal_indices is not None:
        si_row = row + 6
        ws.cell(row=si_row, column=1, value="■ 季節指数").font = SECTION_FONT
        si_row += 1
        month_names = ["1月","2月","3月","4月","5月","6月",
                       "7月","8月","9月","10月","11月","12月"]
        for i, mn in enumerate(month_names):
            c = ws.cell(row=si_row, column=i + 1, value=mn)
            c.font = Font(name=FONT_NAME, bold=True, size=9)
            c.fill = LIGHT_FILL
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER
        si_row += 1
        for i in range(12):
            val = seasonal_indices.get(i + 1, 1.0)
            c = ws.cell(row=si_row, column=i + 1, value=val)
            c.number_format = '0.00'
            c.alignment = Alignment(horizontal="center")
            c.font = Font(name=FONT_NAME, size=10)
            c.border = THIN_BORDER
            if val >= 1.1:
                c.fill = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")
            elif val <= 0.9:
                c.fill = YELLOW_FILL
        usage_start = si_row + 2
    else:
        usage_start = row + 6

    # 使い方
    ws.cell(row=usage_start, column=1, value="■ 使い方").font = SECTION_FONT
    instructions = [
        "1. B5セルの発注係数を変更すると、目標在庫・発注推奨数が自動再計算されます",
        "2.「発注詳細」シートのH列（発注残）に既存の発注残数を入力してください",
        "3. 緊急度（S列）を参考に発注優先度を判断してください",
        f"4. パラメータ: リードタイム={params['lt']}日 / 発注サイクル={params['oc']}日",
    ]
    for i, txt in enumerate(instructions):
        ws.cell(row=usage_start + 1 + i, column=1, value=txt).font = Font(
            name=FONT_NAME, size=9, color="444444"
        )

    _set_col_widths(ws, [20, 12, 20, 10, 10, 10, 10, 10, 10, 10, 10, 10])


# ──────────────────────────────────────────────
# シート2: 発注詳細
# ──────────────────────────────────────────────

def _build_detail_sheet(ws, sku_data, maker_type, params):
    lt = params["lt"]
    oc = params["oc"]
    lt_oc = lt + oc
    ncols = len(DETAIL_HEADERS)

    # ヘッダー行
    for c, hdr in enumerate(DETAIL_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    _set_col_widths(ws, COL_WIDTHS)
    ws.freeze_panes = "A2"

    # セクション分け
    if maker_type == "flexi":
        sections = [
            ("【本体】", [d for d in sku_data if d["category"] == "本体"]),
            ("【アクセサリー】", [d for d in sku_data if d["category"] == "ｱｸｾｻﾘ"]),
        ]
    else:
        sections = [
            ("【通常サイズ】", [d for d in sku_data if d["category"] == "通常"]),
            ("【大容量】", [d for d in sku_data if d["category"] == "大容量"]),
            ("【サンプル】", [d for d in sku_data if d["category"] == "ｻﾝﾌﾟﾙ"]),
        ]

    row = 2
    subtotal_rows = {}
    data_row_ranges = []  # [(start, end)] for each section

    for sec_idx, (sec_name, items) in enumerate(sections):
        if not items:
            continue
        # セクションヘッダー
        ws.cell(row=row, column=1, value=sec_name).font = SECTION_FONT
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = SECTION_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1
        sec_data_start = row

        # データ行
        for item in items:
            _write_data_row(ws, row, item, lt_oc)
            row += 1
        sec_data_end = row - 1
        data_row_ranges.append((sec_data_start, sec_data_end))

        # 小計行
        _write_subtotal_row(ws, row, sec_data_start, sec_data_end, ncols, sec_name.replace("【", "").replace("】", "") + " 小計")
        if sec_idx == 0:
            subtotal_rows["main_subtotal"] = row
        elif sec_idx == 1:
            subtotal_rows["sub_subtotal"] = row
        row += 1
        row += 1  # 空行

    # 合計行
    if data_row_ranges:
        _write_grand_total_row(ws, row, data_row_ranges, ncols)
        subtotal_rows["grand_total"] = row
    else:
        subtotal_rows["grand_total"] = None

    # 条件付き書式 (S列: 緊急度)
    max_row = row
    s_col = get_column_letter(19)  # S列
    range_str = f"{s_col}2:{s_col}{max_row}"
    ws.conditional_formatting.add(range_str, CellIsRule(
        operator="equal", formula=['"緊急"'], fill=RED_FILL
    ))
    ws.conditional_formatting.add(range_str, CellIsRule(
        operator="equal", formula=['"注意"'], fill=YELLOW_FILL
    ))
    ws.conditional_formatting.add(range_str, CellIsRule(
        operator="equal", formula=['"通常"'], fill=GREEN_FILL
    ))

    return subtotal_rows


def _write_data_row(ws, row, item, lt_oc):
    """データ行を書き込む（値 + 数式）"""
    r = row
    # A: SKUコード
    ws.cell(row=r, column=1, value=item["sku_code"]).font = NORMAL_FONT
    # B: 商品名
    ws.cell(row=r, column=2, value=item["product_name"]).font = NORMAL_FONT
    # C: 区分
    ws.cell(row=r, column=3, value=item["category"]).font = NORMAL_FONT
    # D: 現在庫合計
    c = ws.cell(row=r, column=4, value=item["stock_total"])
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_INT
    # E: カインズ
    c = ws.cell(row=r, column=5, value=item["stock_cainz"])
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_INT
    # F: RSL
    c = ws.cell(row=r, column=6, value=item["stock_rsl"])
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_INT
    # G: FBA
    c = ws.cell(row=r, column=7, value=item["stock_fba"])
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_INT
    # H: 発注残 (ユーザー入力用、デフォルト0)
    c = ws.cell(row=r, column=8, value=0)
    c.font = BACKORDER_FONT
    c.fill = BACKORDER_FILL
    c.number_format = NUM_FMT_INT
    # I: 過去12ヶ月販売
    c = ws.cell(row=r, column=9, value=item["sales_12m"])
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_INT
    # J: 月平均(6ヶ月)
    c = ws.cell(row=r, column=10, value=item["avg_6m"])
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_DEC
    # K: 月間予測需要
    c = ws.cell(row=r, column=11, value=item["monthly_forecast"])
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_DEC
    # L: 安全在庫
    c = ws.cell(row=r, column=12, value=item["safety_stock"])
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_DEC

    # M: 目標在庫 = ROUND((K/30*(LT+OC)+L)*シミュレーション!$B$5, 0)
    c = ws.cell(row=r, column=13)
    c.value = f'=ROUND((K{r}/30*{lt_oc}+L{r})*シミュレーション!$B$5,0)'
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_INT
    # N: 発注推奨数 = IF(M-D-H>0, CEILING(M-D-H, Q), 0)  ケース入数が0/1の場合はROUNDUP
    c = ws.cell(row=r, column=14)
    c.value = f'=IF(M{r}-D{r}-H{r}>0,IF(Q{r}>1,CEILING(M{r}-D{r}-H{r},Q{r}),ROUNDUP(M{r}-D{r}-H{r},0)),0)'
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_INT
    # O: 発注ケース数 = IF(N>0, N/Q, 0)
    c = ws.cell(row=r, column=15)
    c.value = f'=IF(AND(N{r}>0,Q{r}>0),N{r}/Q{r},0)'
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_INT
    # P: 発注金額概算 = N*R
    c = ws.cell(row=r, column=16)
    c.value = f'=N{r}*R{r}'
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_YEN
    # Q: ケース入数
    c = ws.cell(row=r, column=17, value=item["case_quantity"])
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_INT
    # R: 仕入単価
    c = ws.cell(row=r, column=18, value=item["unit_cost"])
    c.font = NORMAL_FONT
    c.number_format = NUM_FMT_YEN
    # S: 緊急度 = IF(D=0,"緊急",IF(D<L,"注意","通常"))
    c = ws.cell(row=r, column=19)
    c.value = f'=IF(D{r}=0,"緊急",IF(D{r}<L{r},"注意","通常"))'
    c.font = NORMAL_FONT
    c.alignment = Alignment(horizontal="center")

    # ボーダー
    _apply_border(ws, r, 19)


def _write_subtotal_row(ws, row, data_start, data_end, ncols, label):
    """小計行"""
    ws.cell(row=row, column=1, value=label).font = Font(
        name=FONT_NAME, bold=True, size=10
    )
    # 合計対象列: D(4), N(14), O(15), P(16)
    for col in [4, 14, 15, 16]:
        cl = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f'=SUM({cl}{data_start}:{cl}{data_end})'
        c.font = Font(name=FONT_NAME, bold=True, size=10)
        if col == 16:
            c.number_format = NUM_FMT_YEN
        else:
            c.number_format = NUM_FMT_INT
    for c_idx in range(1, ncols + 1):
        ws.cell(row=row, column=c_idx).fill = SUBTOTAL_FILL
        ws.cell(row=row, column=c_idx).border = THIN_BORDER


def _write_grand_total_row(ws, row, data_row_ranges, ncols):
    """合計行"""
    ws.cell(row=row, column=1, value="合計").font = Font(
        name=FONT_NAME, bold=True, size=11, color=MAIN_COLOR
    )
    for col in [4, 14, 15, 16]:
        cl = get_column_letter(col)
        # 全セクションの範囲を合算
        parts = [f'{cl}{s}:{cl}{e}' for s, e in data_row_ranges]
        formula = '=SUM(' + ','.join(parts) + ')'
        c = ws.cell(row=row, column=col)
        c.value = formula
        c.font = Font(name=FONT_NAME, bold=True, size=11, color=MAIN_COLOR)
        if col == 16:
            c.number_format = NUM_FMT_YEN
        else:
            c.number_format = NUM_FMT_INT
    for c_idx in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c_idx)
        cell.fill = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
        cell.border = Border(
            left=Side(style="thin", color="C0C0C0"),
            right=Side(style="thin", color="C0C0C0"),
            top=Side(style="medium", color=MAIN_COLOR),
            bottom=Side(style="medium", color=MAIN_COLOR),
        )


# ──────────────────────────────────────────────
# シート3: 季節指数
# ──────────────────────────────────────────────

def _build_seasonal_sheet(ws, seasonal_indices):
    ws.sheet_properties.tabColor = "2E86C1"

    ws.merge_cells("A1:M1")
    ws["A1"] = "月別 季節指数"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=14, color=MAIN_COLOR)

    ws["A3"] = "月"
    ws["B3"] = "季節指数"
    ws["C3"] = "判定"
    for c in range(1, 4):
        cell = ws.cell(row=3, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    month_names = ["1月","2月","3月","4月","5月","6月",
                   "7月","8月","9月","10月","11月","12月"]
    for i, mn in enumerate(month_names):
        r = 4 + i
        ws.cell(row=r, column=1, value=mn).font = NORMAL_FONT
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=1).border = THIN_BORDER

        val = seasonal_indices.get(i + 1, 1.0)
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = '0.00'
        c.font = NORMAL_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER

        c3 = ws.cell(row=r, column=3)
        c3.font = NORMAL_FONT
        c3.alignment = Alignment(horizontal="center")
        c3.border = THIN_BORDER
        if val >= 1.1:
            c.fill = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")
            c3.value = "ピーク"
            c3.fill = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")
        elif val <= 0.9:
            c.fill = YELLOW_FILL
            c3.value = "閑散"
            c3.fill = YELLOW_FILL
        else:
            c3.value = "-"

    _set_col_widths(ws, [8, 12, 8])
